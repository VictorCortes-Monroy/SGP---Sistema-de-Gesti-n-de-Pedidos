import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def export_requests_excel(requests: List) -> io.BytesIO:
    """Generate Excel file from a list of Request objects."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"

    headers = ["ID", "Titulo", "Estado", "Monto Total", "Moneda", "Fecha Creacion"]
    ws.append(headers)

    for req in requests:
        ws.append([
            str(req.id),
            req.title,
            req.status.value if hasattr(req.status, "value") else str(req.status),
            float(req.total_amount),
            req.currency,
            req.created_at.strftime("%Y-%m-%d %H:%M") if req.created_at else "",
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_requests_pdf(requests: List) -> io.BytesIO:
    """Generate PDF file from a list of Request objects."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("Reporte de Solicitudes", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["ID", "Titulo", "Estado", "Monto", "Moneda", "Fecha"]]
    for req in requests:
        data.append([
            str(req.id)[:8] + "...",
            (req.title[:30] + "...") if len(req.title) > 30 else req.title,
            req.status.value if hasattr(req.status, "value") else str(req.status),
            f"${float(req.total_amount):,.2f}",
            req.currency,
            req.created_at.strftime("%Y-%m-%d") if req.created_at else "",
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)
    return output


# ── Budget Report Export ──

STATUS_LABELS = {
    "DRAFT": "Borrador",
    "PENDING_TECHNICAL": "Pendiente Tecnico",
    "PENDING_FINANCIAL": "Pendiente Financiero",
    "APPROVED": "Aprobada",
    "REJECTED": "Rechazada",
    "PURCHASING": "En Compra",
    "RECEIVED_PARTIAL": "Recepcion Parcial",
    "RECEIVED_FULL": "Recepcion Total",
    "COMPLETED": "Completada",
    "CANCELLED": "Cancelada",
}

ACTION_LABELS = {
    "CREATED": "Creada",
    "SUBMITTED": "Enviada",
    "APPROVED": "Aprobada",
    "REJECTED": "Rechazada",
    "CANCELLED": "Cancelada",
    "RESUBMITTED": "Reenviada",
    "MARKED_PURCHASING": "En Compra",
    "RECEIVED_PARTIAL": "Recepcion Parcial",
    "RECEIVED_FULL": "Recepcion Total",
    "COMPLETED": "Completada",
}


def export_budget_excel(report_data: dict) -> io.BytesIO:
    """Generate Excel file from budget report data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    group_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.append([f"Reporte de Presupuestos - Año {report_data['year']}"])
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    # Grand totals
    ws.append(["Resumen General"])
    ws["A3"].font = Font(bold=True, size=12)
    ws.append(["Total", "Reservado", "Ejecutado", "Disponible"])
    for cell in ws[4]:
        if cell.value:
            cell.font = Font(bold=True)
    ws.append([
        float(report_data["grand_total"]),
        float(report_data["grand_reserved"]),
        float(report_data["grand_executed"]),
        float(report_data["grand_available"]),
    ])
    for cell in ws[5]:
        if cell.value is not None:
            cell.number_format = '#,##0.00'
    ws.append([])

    # Per-company groups
    headers = ["Centro de Costo", "Codigo", "Total", "Reservado", "Ejecutado", "Disponible", "Utilizacion %"]
    for group in report_data["groups"]:
        # Company header
        ws.append([f"Empresa: {group['company_name']}"])
        row_num = ws.max_row
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).fill = group_fill
            ws.cell(row=row_num, column=col).font = Font(bold=True, size=11)

        # Column headers
        ws.append(headers)
        row_num = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for item in group["budgets"]:
            ws.append([
                item["cost_center_name"],
                item["cost_center_code"],
                float(item["total_amount"]),
                float(item["reserved_amount"]),
                float(item["executed_amount"]),
                float(item["available_amount"]),
                round(item["utilization_pct"], 1),
            ])
            row_num = ws.max_row
            for col in range(3, 7):
                ws.cell(row=row_num, column=col).number_format = '#,##0.00'
            ws.cell(row=row_num, column=7).number_format = '0.0"%"'
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = thin_border

        ws.append([])

    # Auto-width (skip merged cells)
    for col in ws.columns:
        cells = [c for c in col if hasattr(c, 'column_letter')]
        if not cells:
            continue
        max_length = max((len(str(cell.value or "")) for cell in cells), default=10)
        ws.column_dimensions[cells[0].column_letter].width = min(max_length + 2, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_budget_pdf(report_data: dict) -> io.BytesIO:
    """Generate PDF file from budget report data."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Reporte de Presupuestos - Año {report_data['year']}", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Grand totals
    summary_data = [
        ["Total", "Reservado", "Ejecutado", "Disponible"],
        [
            f"${float(report_data['grand_total']):,.2f}",
            f"${float(report_data['grand_reserved']):,.2f}",
            f"${float(report_data['grand_executed']):,.2f}",
            f"${float(report_data['grand_available']):,.2f}",
        ],
    ]
    t = Table(summary_data, colWidths=[150, 150, 150, 150])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    for group in report_data["groups"]:
        elements.append(Paragraph(f"Empresa: {group['company_name']}", styles["Heading2"]))
        elements.append(Spacer(1, 6))

        data = [["Centro Costo", "Codigo", "Total", "Reservado", "Ejecutado", "Disponible", "Util. %"]]
        for item in group["budgets"]:
            data.append([
                item["cost_center_name"][:20],
                item["cost_center_code"],
                f"${float(item['total_amount']):,.2f}",
                f"${float(item['reserved_amount']):,.2f}",
                f"${float(item['executed_amount']):,.2f}",
                f"${float(item['available_amount']):,.2f}",
                f"{item['utilization_pct']:.1f}%",
            ])

        table = Table(data, colWidths=[100, 70, 90, 90, 90, 90, 60])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 16))

    doc.build(elements)
    output.seek(0)
    return output


# ── Audit Log Export ──

def export_audit_excel(logs: List) -> io.BytesIO:
    """Generate Excel file from audit log entries."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.append(["Reporte de Auditoria"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["Fecha/Hora", "Actor", "Rol", "Accion", "Solicitud", "De Estado", "A Estado", "Comentario", "IP"]
    ws.append(headers)
    row_num = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for log in logs:
        ts = log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else ""
        ws.append([
            ts,
            log.actor_name or "",
            log.actor_role or "",
            ACTION_LABELS.get(log.action, log.action),
            log.request_title or str(log.request_id)[:8],
            STATUS_LABELS.get(log.from_status, log.from_status or ""),
            STATUS_LABELS.get(log.to_status, log.to_status or ""),
            log.comment or "",
            log.ip_address or "",
        ])
        row_num = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = thin_border

    for col in ws.columns:
        cells = [c for c in col if hasattr(c, 'column_letter')]
        if not cells:
            continue
        max_length = max((len(str(cell.value or "")) for cell in cells), default=10)
        ws.column_dimensions[cells[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_audit_pdf(logs: List) -> io.BytesIO:
    """Generate PDF file from audit log entries."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Reporte de Auditoria", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["Fecha", "Actor", "Rol", "Accion", "Solicitud", "Transicion", "Comentario"]]
    for log in logs:
        ts = log.timestamp.strftime("%Y-%m-%d %H:%M") if log.timestamp else ""
        from_s = STATUS_LABELS.get(log.from_status, log.from_status or "-")
        to_s = STATUS_LABELS.get(log.to_status, log.to_status or "-")
        transition = f"{from_s} → {to_s}" if log.from_status else to_s
        data.append([
            ts,
            (log.actor_name or "")[:15],
            (log.actor_role or "")[:15],
            ACTION_LABELS.get(log.action, log.action),
            (log.request_title or str(log.request_id)[:8])[:20],
            transition[:25],
            (log.comment or "")[:30],
        ])

    table = Table(data, colWidths=[80, 80, 80, 70, 100, 110, 120])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)
    return output


# ── Maintenance Request Export ──

def export_maint_requests_excel(requests: List) -> io.BytesIO:
    """Generate Excel file from a list of MaintRequest objects."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenciones"

    headers = ["ID", "Código", "Equipo", "Tipo", "Estado", "Fecha Planificada", "Fecha Creación"]
    ws.append(headers)

    for req in requests:
        ws.append([
            str(req.id),
            req.code,
            req.equipment.name if hasattr(req, "equipment") and req.equipment else str(req.equipment_id),
            req.maintenance_type.value if hasattr(req.maintenance_type, "value") else str(req.maintenance_type),
            req.status.value if hasattr(req.status, "value") else str(req.status),
            req.planned_date.strftime("%Y-%m-%d") if req.planned_date else "",
            req.created_at.strftime("%Y-%m-%d %H:%M") if req.created_at else "",
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_maint_requests_pdf(requests: List) -> io.BytesIO:
    """Generate PDF file from a list of MaintRequest objects."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("Reporte de Mantenciones Preventivas", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["Código", "Equipo", "Tipo", "Estado", "Fecha Planificada"]]
    for req in requests:
        equip_name = req.equipment.name if hasattr(req, "equipment") and req.equipment else str(req.equipment_id)
        data.append([
            req.code,
            (equip_name[:30] + "...") if len(equip_name) > 30 else equip_name,
            req.maintenance_type.value if hasattr(req.maintenance_type, "value") else str(req.maintenance_type),
            req.status.value if hasattr(req.status, "value") else str(req.status),
            req.planned_date.strftime("%Y-%m-%d") if req.planned_date else "",
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)
    return output
